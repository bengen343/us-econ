"""Manheim Used Vehicle Value Index (Cox Automotive) collector.

Manheim's wholesale used-vehicle index is the best-documented free leading
indicator for CPI used cars & trucks (wholesale moves flow to retail with a
~1-2 month lag), and the full month's value is published on the 5th business
day after month-end — i.e. fully observed for month M before M's CPI release.

Each monthly release is a "trends" post on coxautoinc.com that links a
spreadsheet carrying the ENTIRE monthly history (Jan-1997 onward: SA index, SA
and NSA dollar prices, seasonal factors). So every run re-pulls the complete
series and appends it vintage-stamped — history comes free, revisions (e.g.
annual seasonal-factor re-estimation) are preserved, and downstream takes the
latest vintage per (measure, observation_month).

Discovery: the post URL is templated by month name; the run tries the most
recent month first and falls back one month (release lag means month M's post
appears ~5 business days into M+1). A 404 on both is "not published yet" —
zero rows, no error.
"""

import io
import logging
import re
from datetime import date

import httpx
import openpyxl
from google.cloud import bigquery

from collectors.common import LoadSpec, Settings
from collectors.common.http import client, with_retries

_log = logging.getLogger(__name__)

TABLE = "manheim_used_vehicles.value_index"

PAGE_URL = (
    "https://www.coxautoinc.com/insights/"
    "manheim-used-vehicle-value-index-{month}-{year}-trends/"
)
_XLSX_RE = re.compile(r'href="(https://www\.coxautoinc\.com/[^"]+\.xlsx)"')

# Some Cox endpoints reject the default collector User-Agent; send a browser UA.
_BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

# DATA sheet layout: col A = month, then the four level columns we keep (the
# remaining columns are % changes, derivable downstream).
_MEASURES: list[tuple[int, str, str]] = [
    (1, "index_sa", "index (1997-01 = 100)"),
    (2, "price_sa", "USD"),
    (3, "seasonal_factor", "factor"),
    (4, "price_nsa", "USD"),
]

SCHEMA: list[bigquery.SchemaField] = [
    bigquery.SchemaField("measure", "STRING", mode="REQUIRED"),
    bigquery.SchemaField("observation_month", "DATE", mode="REQUIRED"),
    bigquery.SchemaField("value", "FLOAT64"),
    bigquery.SchemaField("units", "STRING"),
    bigquery.SchemaField("reference_month", "DATE", mode="REQUIRED"),  # release's named month
    bigquery.SchemaField("source_url", "STRING"),
]


def collect(settings: Settings) -> LoadSpec:
    rows: list[dict] = []
    with client() as http:
        for reference_month in _candidate_months(date.today()):
            page = _get_page(http, reference_month)
            if page is None:
                continue
            page_url, html = page
            match = _XLSX_RE.search(html)
            if match is None:
                raise RuntimeError(f"no xlsx data link found on {page_url}")
            workbook = _get_bytes(http, match.group(1))
            rows = _parse_workbook(workbook, reference_month, match.group(1))
            _log.info(
                "Manheim index parsed",
                extra={
                    "extras": {
                        "reference_month": reference_month.isoformat(),
                        "row_count": len(rows),
                        "page": page_url,
                        "xlsx": match.group(1),
                    }
                },
            )
            break
    if not rows:
        _log.info("no Manheim release page found (not published yet); skipping load")
    return LoadSpec(table=TABLE, schema=SCHEMA, rows=rows)


def _candidate_months(today: date) -> list[date]:
    """Most recent possibly-published reference month first, then one older.

    Month M's full-month release lands ~5 business days into M+1, so the
    freshest possible reference month is always last month; the fallback covers
    a slipped run early in the window.
    """
    first = date(today.year, today.month, 1)
    last_month = _prior_month(first)
    return [last_month, _prior_month(last_month)]


def _prior_month(d: date) -> date:
    return date(d.year - 1, 12, 1) if d.month == 1 else date(d.year, d.month - 1, 1)


def _get_page(http: httpx.Client, reference_month: date) -> tuple[str, str] | None:
    url = PAGE_URL.format(
        month=reference_month.strftime("%B").lower(), year=reference_month.year
    )

    def call() -> httpx.Response:
        response = http.get(
            url, headers={"User-Agent": _BROWSER_UA, "Accept": "text/html"}
        )
        if response.status_code == 404:
            return response  # not published yet -- caller falls back
        response.raise_for_status()
        return response

    response = with_retries(call)
    if response.status_code == 404:
        return None
    return url, response.text


def _get_bytes(http: httpx.Client, url: str) -> bytes:
    def call() -> bytes:
        response = http.get(url, headers={"User-Agent": _BROWSER_UA})
        response.raise_for_status()
        return response.content

    return with_retries(call)


def _parse_workbook(payload: bytes, reference_month: date, source_url: str) -> list[dict]:
    """Long rows from the DATA sheet (full monthly history, one row per measure).

    data_only=True returns cached values for the handful of formula cells the
    workbook carries; a cell that is not numeric is skipped (NULL downstream).
    """
    wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    if "DATA" not in wb.sheetnames:
        raise RuntimeError(f"no DATA sheet in workbook from {source_url}")
    rows: list[dict] = []
    for record in wb["DATA"].iter_rows(values_only=True):
        month = record[0]
        if not hasattr(month, "year") or not hasattr(month, "month"):
            continue  # header / spacer rows
        for col, measure, units in _MEASURES:
            value = record[col] if col < len(record) else None
            if not isinstance(value, (int, float)):
                continue
            rows.append(
                {
                    "measure": measure,
                    "observation_month": date(month.year, month.month, 1).isoformat(),
                    "value": float(value),
                    "units": units,
                    "reference_month": reference_month.isoformat(),
                    "source_url": source_url,
                }
            )
    latest = max((r["observation_month"] for r in rows), default=None)
    if latest != reference_month.isoformat():
        raise RuntimeError(
            f"workbook history ends {latest}, expected the release's named month "
            f"{reference_month} -- layout changed?"
        )
    return rows
